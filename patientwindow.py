import customtkinter
import tkcalendar
from tkcalendar import DateEntry
from PIL import Image, ImageTk
from tkinter import messagebox, filedialog
import tkinter
from datetime import datetime
from sarandb import Database
from resource_1 import resource_path
import openpyxl
import os
from searchwindow import SearchPatient
from pdfgeneration import create_pdf

class PatientWindow(customtkinter.CTk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.configure_settings()
        self.db = Database()
        self.photo_path = None  # Initialize photo_path here

    def configure_settings(self):
        customtkinter.set_appearance_mode("dark")
        customtkinter.set_default_color_theme("blue")
        self.create_window()

    def create_window(self):
        self.title("Patient Information")
        self.after(0, lambda: self.state('zoomed'))

        self.img = customtkinter.CTkImage(light_image=Image.open(resource_path(r'C:\Users\soman\prefinal_saran\icons\logo.png')), size=(100, 100))
        self.img_label = customtkinter.CTkLabel(self, image=self.img, text=" ")
        self.img_label.pack()

        self.saran_label = customtkinter.CTkLabel(self, text="SARAN NASHA MUKTI KENDRA", corner_radius=5, fg_color="#151616", text_color="#E9410E", font=("Arial", 20))
        self.saran_label.place(x=520, y=125)

        self.date_label = customtkinter.CTkLabel(self, text=datetime.now().strftime("%d-%m-%Y"), corner_radius=5, fg_color="#4B4D50", text_color="#E6164A", font=("Arial", 12))
        self.date_label.place(x=900, y=125)

        # Time label
        self.time_label = customtkinter.CTkLabel(self, text=datetime.now().strftime("%H:%M:%S"), corner_radius=5, fg_color="#4B4D50", text_color="#E6164A", font=("Arial", 12))
        self.time_label.place(x=1050, y=125)
        self.update_time()

        self.patient_registration_label = customtkinter.CTkLabel(self, text="Registration Number", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.patient_registration_label.place(x=20, y=200)

        self.patient_registration_entry = customtkinter.CTkEntry(self, placeholder_text="Registration Number", width=200)
        self.patient_registration_entry.place(x=200, y=200)

        self.patient_name_label = customtkinter.CTkLabel(self, text="Patient Name", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.patient_name_label.place(x=500, y=200)

        self.patient_name_entry = customtkinter.CTkEntry(self, placeholder_text="Patient Name", width=200)
        self.patient_name_entry.place(x=640, y=200)

        # Frame to Upload Patient Photo
        self.photo_frame = customtkinter.CTkFrame(self, height=200, width=200, corner_radius=10, fg_color="#96CFE9", border_color="yellow")
        self.photo_frame.place(x=1000, y=200)

        # Upload Button to upload patient photo.. This should be just below the frame
        self.upload_button = customtkinter.CTkButton(self, text="Upload Photo", fg_color="#4B4D50", hover_color="#10A2E6", command=self.upload_pic)
        self.upload_button.place(x=1025, y=450)

        self.patient_address_label = customtkinter.CTkLabel(self, text="Address", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.patient_address_label.place(x=20, y=300)

        self.patient_address_entry = customtkinter.CTkEntry(self, placeholder_text="Patient Address", width=250)
        self.patient_address_entry.place(x=200, y=300)

        self.date_of_joining_label = customtkinter.CTkLabel(self, text="Date of Joining", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.date_of_joining_label.place(x=500, y=300)

        # Date of Joining Entry (Calendar)
        # self.date_of_joining_entry = DateEntry(self, width=10, foreground='#FD0303', borderwidth=2)
        # self.date_of_joining_entry.place(x=640, y=300)

        # TKcalendar not working when creating EXE so Just add an Entry Box
        self.day_entry = customtkinter.CTkEntry(self, width=40, corner_radius=5, placeholder_text="DD")
        self.day_entry.place(x=640, y=300)

        self.month_entry = customtkinter.CTkEntry(self, width=40, corner_radius=5, placeholder_text="MM")
        self.month_entry.place(x=690, y=300)

        self.year_entry = customtkinter.CTkEntry(self, width=60, corner_radius=5, placeholder_text="YYYY")
        self.year_entry.place(x=740, y=300)

        # Addiction Type
        self.addiction_type_label = customtkinter.CTkLabel(self, text="Addiction Type", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.addiction_type_label.place(x=20, y=400)

        self.addiction_type_combo = customtkinter.CTkComboBox(self, width=200, corner_radius=5, fg_color="#10A2E6", text_color="#000", values=["Alcohol", "Brown Sugar", "Smack", "Gambling", "Weed (Ganja)", "Nicotine", "Whitener", "Cocaine", "LSD", "MDMA", "Sleeping Pills"])
        self.addiction_type_combo.place(x=200, y=400)

        self.addiction_duration_label = customtkinter.CTkLabel(self, text="Addiction Duration", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.addiction_duration_label.place(x=500, y=400)

        self.addiction_duration = customtkinter.CTkComboBox(self, width=100, corner_radius=5, fg_color="#10A2E6", text_color="#000", values=["1 Year", "2 Year", "3 Year", "4 Year", "5 Year", "> 5 Year", "> 10 Year"])
        self.addiction_duration.place(x=650, y=400)

        # Gender
        self.radiovar = tkinter.IntVar(value=0)

        self.gender_label = customtkinter.CTkLabel(self, text="Gender", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.gender_label.place(x=20, y=500)

        self.gender_radio_male = customtkinter.CTkRadioButton(self, text="Male", variable=self.radiovar, value=1)
        self.gender_radio_male.place(x=200, y=500)

        self.gender_radio_female = customtkinter.CTkRadioButton(self, text="Female", variable=self.radiovar, value=2)
        self.gender_radio_female.place(x=300, y=500)

        # Room Types
        self.room_var = tkinter.StringVar()
        self.room_types = customtkinter.CTkComboBox(self, width=200, corner_radius=5, fg_color="#10A2E6", text_color="#000", values=["AC-Room", "AC-Hall", "Non-AC"], command=self.update_charges)
        self.room_types.place(x=500, y=500)

        # Rehab Charges
        self.monthly_charge_label = customtkinter.CTkLabel(self, text="Monthly Charge", corner_radius=5, fg_color="#10A2E6", text_color="#0C0D0E", font=("Arial", 14))
        self.monthly_charge_label.place(x=750, y=500)
        self.monthly_charge_entry = customtkinter.CTkEntry(self, state="readonly")
        self.monthly_charge_entry.place(x=900, y=500)

        # Submit Button
        self.submit_button = customtkinter.CTkButton(self, text="Submit", width=200, corner_radius=6, border_width=1, border_color="blue", fg_color="#000", command=self.collect_data)
        self.submit_button.place(x=10, y=550)

        # Exit Button
        self.exit_button = customtkinter.CTkButton(self, text="Exit", width=200, corner_radius=6, border_width=1, border_color="blue", fg_color="#000", command=self.quit)
        self.exit_button.place(x=250, y=550)

        # New Window to Search for Patients
        self.search_button = customtkinter.CTkButton(self, text="Search Patient", width=200, corner_radius=6, border_width=1, border_color="blue", fg_color="#000", command=self.search)
        self.search_button.place(x=550, y=550)

        self.pay_button = customtkinter.CTkButton(self, text="Payment", width=200, corner_radius=6, border_width=1, border_color="blue", fg_color="#000", command=self.pay)
        self.pay_button.place(x=0, y=650)

        self.pdf_button = customtkinter.CTkButton(self, text="Generate PDF", width=200, corner_radius=6, border_width=1, border_color="blue", fg_color="#000", command=self.pdf)
        self.pdf_button.place(x=250, y=650)

    def pay(self):
        pass

    def pdf(self):
        self.collect_data()

    def update_time(self):
        current_time = datetime.now().strftime("%H:%M:%S")
        self.time_label.configure(text=current_time)
        self.after(1000, self.update_time)  # Update time every second

    def search(self):
        s = SearchPatient()
        s.search_patient()
        s.mainloop()

    def upload_pic(self):
        file_path = filedialog.askopenfilename()
        print(file_path)
        try:
            if file_path:
                self.photo_path = file_path  # Update the photo_path attribute
                self.patient_photo = customtkinter.CTkImage(light_image=Image.open(file_path), size=(200, 200))
                for widget in self.photo_frame.winfo_children():
                    widget.destroy()
                self.patient_photo_label = customtkinter.CTkLabel(self.photo_frame, image=self.patient_photo, text=" ")
                self.patient_photo_label.pack()
        except Exception as e:
            messagebox.showerror("Error In uploading Photo", message=str(e))

    def update_charges(self, value):
        if value == "AC-Room":
            charge = 20000
        elif value == "AC-Hall":
            charge = 15000
        else:
            charge = 12500
        self.monthly_charge_entry.configure(state="normal")
        self.monthly_charge_entry.delete(0, tkinter.END)
        self.monthly_charge_entry.insert(0, str(charge))
        self.monthly_charge_entry.configure(state="readonly")

    def collect_data(self):
        registration_number = self.patient_registration_entry.get()
        patient_name = self.patient_name_entry.get()
        patient_address = self.patient_address_entry.get()
        day = self.day_entry.get()
        month = self.month_entry.get()
        year = self.year_entry.get()
        doj = day + "/" + month + "/" + year
        addiction_type = self.addiction_type_combo.get()
        gender = self.radiovar.get()
        sex = 'Male' if gender == 1 else 'Female'
        duration = self.addiction_duration.get()
        room_type = self.room_types.get()
        monthly_charge = self.monthly_charge_entry.get()
        photo_path = self.photo_path  # Use the photo_path attribute

        patient_data = (
            registration_number, patient_name, patient_address, doj, addiction_type, sex, duration, room_type, monthly_charge, photo_path
        )
        try:
            self.db.insert_patient_data(patient_data)
            self.save_workbook()
            print("Creating PDF ")
            create_pdf(reg_no=registration_number, name=patient_name, address=patient_address, doj=doj, addiction_type=addiction_type, gender=sex, duration=duration, room_type=room_type, monthly_charge=monthly_charge, photo_path=photo_path)
            self.clear_fields()
            messagebox.showinfo(title="Success", message="Inserted Patient Data AND... Making Entry in Excel")
        except Exception as e:
            messagebox.showerror(title="Error", message=str(e))

    def clear_fields(self):
        self.patient_registration_entry.delete(0, tkinter.END)
        self.patient_name_entry.delete(0, tkinter.END)
        self.patient_address_entry.delete(0, tkinter.END)
        self.day_entry.delete(0, tkinter.END)
        self.month_entry.delete(0, tkinter.END)
        self.year_entry.delete(0, tkinter.END)
        self.addiction_type_combo.set('')
        self.radiovar.set(0)
        self.addiction_duration.set('')
        self.room_types.set('')
        self.monthly_charge_entry.configure(state="normal")
        self.monthly_charge_entry.delete(0, tkinter.END)
        self.monthly_charge_entry.configure(state="readonly")
        self.photo_path = None

    def save_workbook(self):
        file_path = 'patient_info.xlsx'
        file_exists = os.path.isfile(file_path)

        if file_exists:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet['A1'] = "Registration Number"
            sheet['B1'] = "Patient Name"
            sheet['C1'] = "Address"
            sheet['D1'] = "Date of Joining"
            sheet['E1'] = "Addiction Type"
            sheet['F1'] = "Gender"
            sheet['G1'] = "Addiction Duration"
            sheet['H1'] = "Room Type"
            sheet['I1'] = "Monthly Charge"
            sheet['J1'] = "Photo Path"

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=self.patient_registration_entry.get())
        sheet.cell(row=next_row, column=2, value=self.patient_name_entry.get())
        sheet.cell(row=next_row, column=3, value=self.patient_address_entry.get())
        sheet.cell(row=next_row, column=4, value=self.day_entry.get() + "/" + self.month_entry.get() + "/" + self.year_entry.get())
        sheet.cell(row=next_row, column=5, value=self.addiction_type_combo.get())
        sheet.cell(row=next_row, column=6, value='Male' if self.radiovar.get() == 1 else 'Female')
        sheet.cell(row=next_row, column=7, value=self.addiction_duration.get())
        sheet.cell(row=next_row, column=8, value=self.room_types.get())
        sheet.cell(row=next_row, column=9, value=self.monthly_charge_entry.get())
        sheet.cell(row=next_row, column=10, value=getattr(self, 'photo_path', None))

        wb.save(file_path)

p = PatientWindow()
p.mainloop()
